"""
05_build_excel.py
Builds Workforce_Analytics_Report.xlsx with:
  1. Raw Data - Departments   (input data, would come from HRIS export)
  2. Raw Data - Recruitment   (input data, would come from ATS export)
  3. Daily Headcount Report   (formulas: SUMIFS/INDEX-MATCH off raw data)
  4. Recruitment Funnel       (formulas: conversion rates by dept, latest quarter)
  5. Headcount Forecast       (what-if model: assumptions -> hires needed)
All computed values are Excel formulas, not hardcoded Python results.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

DATA = "/home/claude/project/data"
OUTFILE = "/home/claude/project/Workforce_Analytics_Report.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E5F")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=9, color="666666")
INPUT_FONT = Font(name=FONT, color="0000FF")
LABEL_FONT = Font(name=FONT, bold=True)
NORMAL_FONT = Font(name=FONT, size=10)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# =================================================================
# SHEET 1: Raw Data - Departments
# =================================================================
ws = wb.active
ws.title = "Raw_Departments"
dept_df = pd.read_csv(f"{DATA}/departments.csv")
emp_df = pd.read_csv(f"{DATA}/employees.csv")

# aggregate current headcount / exits per dept for raw pull
agg = emp_df.groupby("dept_id").agg(
    current_headcount=("is_active", "sum"),
    total_ever_hired=("employee_id", "count"),
    exits=("is_active", lambda s: (~s).sum()),
    avg_engagement=("engagement_score", "mean"),
    avg_salary=("annual_salary_inr", "mean"),
).reset_index()
raw_dept = dept_df.merge(agg, on="dept_id")
raw_dept["avg_engagement"] = raw_dept["avg_engagement"].round(2)
raw_dept["avg_salary"] = raw_dept["avg_salary"].round(0)

ws["A1"] = "RAW DATA — Department Headcount Extract (source: HRIS export, as of 2026-07-01)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "This tab simulates a raw system export. Reporting tabs reference these cells via formulas."
ws["A2"].font = SUBTITLE_FONT

headers = list(raw_dept.columns)
header_row = 4
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=j, value=h.replace("_", " ").title())
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

for i, row in raw_dept.iterrows():
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row + 1 + i, column=j, value=row[h])
        c.font = NORMAL_FONT
        c.border = BORDER

for j, h in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(j)].width = 20

DEPT_LAST_ROW = header_row + len(raw_dept)

# =================================================================
# SHEET 2: Raw Data - Recruitment Funnel
# =================================================================
ws2 = wb.create_sheet("Raw_Recruitment")
funnel_df = pd.read_csv(f"{DATA}/recruitment_funnel.csv")

ws2["A1"] = "RAW DATA — Recruitment Funnel Extract (source: ATS export, by quarter & department)"
ws2["A1"].font = TITLE_FONT

f_headers = list(funnel_df.columns)
for j, h in enumerate(f_headers, start=1):
    c = ws2.cell(row=3, column=j, value=h.replace("_", " ").title())
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

for i, row in funnel_df.iterrows():
    for j, h in enumerate(f_headers, start=1):
        c = ws2.cell(row=4 + i, column=j, value=row[h])
        c.font = NORMAL_FONT
        c.border = BORDER

for j, h in enumerate(f_headers, start=1):
    ws2.column_dimensions[get_column_letter(j)].width = 14

FUNNEL_LAST_ROW = 3 + len(funnel_df)
LATEST_QUARTER = funnel_df["quarter"].max()

# =================================================================
# SHEET 3: Daily Headcount Report  (formulas off Raw_Departments)
# =================================================================
ws3 = wb.create_sheet("Daily_Headcount_Report")
wb.move_sheet("Daily_Headcount_Report", offset=-2)

ws3["A1"] = "Daily Headcount Report"
ws3["A1"].font = TITLE_FONT
ws3["A2"] = "=\"Generated: \" & TEXT(TODAY(), \"dd-mmm-yyyy\") & \"  |  Refreshes automatically when Raw_Departments updates\""
ws3["A2"].font = SUBTITLE_FONT

rep_headers = ["Department", "Target Headcount 2026", "Current Headcount",
                "Gap to Target", "% to Target", "Attrition % (All-Time)", "Status"]
for j, h in enumerate(rep_headers, start=1):
    c = ws3.cell(row=4, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

n_depts = len(raw_dept)
for i in range(n_depts):
    r = 5 + i
    src_r = header_row + 1 + i  # matching row in Raw_Departments
    ws3.cell(row=r, column=1, value=f"=Raw_Departments!B{src_r}").font = NORMAL_FONT
    ws3.cell(row=r, column=2, value=f"=Raw_Departments!C{src_r}").font = NORMAL_FONT
    ws3.cell(row=r, column=3, value=f"=Raw_Departments!D{src_r}").font = NORMAL_FONT
    ws3.cell(row=r, column=4, value=f"=B{r}-C{r}").font = NORMAL_FONT
    ws3.cell(row=r, column=5, value=f"=C{r}/B{r}").font = NORMAL_FONT
    ws3.cell(row=r, column=5).number_format = "0.0%"
    ws3.cell(row=r, column=6, value=f"=Raw_Departments!F{src_r}/Raw_Departments!E{src_r}").font = NORMAL_FONT
    ws3.cell(row=r, column=6).number_format = "0.0%"
    ws3.cell(row=r, column=7,
             value=f'=IF(D{r}>50,"Hiring Gap - High Priority",IF(D{r}>0,"On Track","At/Above Target"))').font = NORMAL_FONT
    for col in range(1, 8):
        ws3.cell(row=r, column=col).border = BORDER

TOT_ROW = 5 + n_depts
ws3.cell(row=TOT_ROW, column=1, value="TOTAL").font = LABEL_FONT
ws3.cell(row=TOT_ROW, column=2, value=f"=SUM(B5:B{TOT_ROW-1})").font = LABEL_FONT
ws3.cell(row=TOT_ROW, column=3, value=f"=SUM(C5:C{TOT_ROW-1})").font = LABEL_FONT
ws3.cell(row=TOT_ROW, column=4, value=f"=SUM(D5:D{TOT_ROW-1})").font = LABEL_FONT
ws3.cell(row=TOT_ROW, column=5, value=f"=C{TOT_ROW}/B{TOT_ROW}").font = LABEL_FONT
ws3.cell(row=TOT_ROW, column=5).number_format = "0.0%"
ws3.cell(row=TOT_ROW, column=6, value=f"=AVERAGE(F5:F{TOT_ROW-1})").font = LABEL_FONT
ws3.cell(row=TOT_ROW, column=6).number_format = "0.0%"

# conditional formatting: highlight big hiring gaps
ws3.conditional_formatting.add(
    f"D5:D{TOT_ROW-1}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                    end_type="max", end_color="F8696B")
)

for j, w in enumerate([18, 18, 16, 14, 12, 16, 20], start=1):
    ws3.column_dimensions[get_column_letter(j)].width = w

# =================================================================
# SHEET 4: Recruitment Funnel Report (formulas off Raw_Recruitment)
# =================================================================
ws4 = wb.create_sheet("Recruitment_Funnel_Report")

ws4["A1"] = "Recruitment Funnel Report — Latest Quarter"
ws4["A1"].font = TITLE_FONT
ws4["A2"] = f"=\"Quarter: \" & \"{LATEST_QUARTER}\""
ws4["A2"].font = SUBTITLE_FONT

fr_headers = ["Department", "Applied", "Screened", "Interviewed", "Offered", "Hired",
              "Applied to Screen %", "Screen to Interview %", "Interview to Offer %",
              "Offer to Hire %", "Overall Conversion %"]
for j, h in enumerate(fr_headers, start=1):
    c = ws4.cell(row=4, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

depts_list = list(raw_dept["dept_name"])
for i, dname in enumerate(depts_list):
    r = 5 + i
    # SUMIFS pulling from Raw_Recruitment filtered by dept + latest quarter
    ws4.cell(row=r, column=1, value=dname).font = NORMAL_FONT
    ws4.cell(row=r, column=2,
             value=f'=SUMIFS(Raw_Recruitment!D:D,Raw_Recruitment!C:C,A{r},Raw_Recruitment!A:A,$A$2)').font = NORMAL_FONT
    ws4.cell(row=r, column=3,
             value=f'=SUMIFS(Raw_Recruitment!E:E,Raw_Recruitment!C:C,A{r},Raw_Recruitment!A:A,$A$2)').font = NORMAL_FONT
    ws4.cell(row=r, column=4,
             value=f'=SUMIFS(Raw_Recruitment!F:F,Raw_Recruitment!C:C,A{r},Raw_Recruitment!A:A,$A$2)').font = NORMAL_FONT
    ws4.cell(row=r, column=5,
             value=f'=SUMIFS(Raw_Recruitment!G:G,Raw_Recruitment!C:C,A{r},Raw_Recruitment!A:A,$A$2)').font = NORMAL_FONT
    ws4.cell(row=r, column=6,
             value=f'=SUMIFS(Raw_Recruitment!H:H,Raw_Recruitment!C:C,A{r},Raw_Recruitment!A:A,$A$2)').font = NORMAL_FONT
    ws4.cell(row=r, column=7, value=f"=IFERROR(C{r}/B{r},0)").number_format = "0.0%"
    ws4.cell(row=r, column=8, value=f"=IFERROR(D{r}/C{r},0)").number_format = "0.0%"
    ws4.cell(row=r, column=9, value=f"=IFERROR(E{r}/D{r},0)").number_format = "0.0%"
    ws4.cell(row=r, column=10, value=f"=IFERROR(F{r}/E{r},0)").number_format = "0.0%"
    ws4.cell(row=r, column=11, value=f"=IFERROR(F{r}/B{r},0)").number_format = "0.0%"
    for col in range(1, 12):
        ws4.cell(row=r, column=col).border = BORDER

# fix A2 formula reference (needs to hold quarter string, referenced by SUMIFS above)
ws4["A2"] = LATEST_QUARTER
ws4["A2"].font = Font(name=FONT, italic=True, size=9, color="0000FF")
ws4["A3"] = "Quarter filter cell above (A2) — change it to any quarter in Raw_Recruitment to refresh this report."
ws4["A3"].font = SUBTITLE_FONT
ws4["A2"].fill = YELLOW_FILL

for j, w in enumerate([18, 10, 10, 12, 10, 10, 14, 16, 14, 12, 14], start=1):
    ws4.column_dimensions[get_column_letter(j)].width = w

# =================================================================
# SHEET 5: Headcount Forecast (what-if model)
# =================================================================
ws5 = wb.create_sheet("Headcount_Forecast")

ws5["A1"] = "Headcount Forecast — What-If Model"
ws5["A1"].font = TITLE_FONT
ws5["A2"] = "Yellow cells are adjustable assumptions. All results recalculate automatically."
ws5["A2"].font = SUBTITLE_FONT

ws5["A4"] = "Assumptions"
ws5["A4"].font = LABEL_FONT
ws5["A5"] = "Planning horizon (months)"
ws5["B5"] = 6
ws5["B5"].fill = YELLOW_FILL
ws5["B5"].font = INPUT_FONT
ws5["A6"] = "Expected monthly attrition rate (company-wide)"
ws5["B6"] = 0.012
ws5["B6"].fill = YELLOW_FILL
ws5["B6"].font = INPUT_FONT
ws5["B6"].number_format = "0.0%"
ws5["A7"] = "Average hire-to-offer lead time (weeks)"
ws5["B7"] = 8
ws5["B7"].fill = YELLOW_FILL
ws5["B7"].font = INPUT_FONT
ws5["A8"] = "Average offer-to-hire conversion rate"
ws5["B8"] = "=Recruitment_Funnel_Report!K9"  # overall conversion avg placeholder, fixed below
ws5["B8"].number_format = "0.0%"

for cell in ["A5", "A6", "A7", "A8"]:
    ws5[cell].font = NORMAL_FONT

ws5["A10"] = "Department Forecast"
ws5["A10"].font = LABEL_FONT

fc_headers = ["Department", "Current Headcount", "Target Headcount", "Net New Hires Needed",
              "Expected Attrition (Horizon)", "Total Hires Required",
              "Applications Needed (at current conversion %)"]
for j, h in enumerate(fc_headers, start=1):
    c = ws5.cell(row=11, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for i, dname in enumerate(depts_list):
    r = 12 + i
    src_r = 5 + i  # matching row in Daily_Headcount_Report
    ws5.cell(row=r, column=1, value=f"=Daily_Headcount_Report!A{src_r}").font = NORMAL_FONT
    ws5.cell(row=r, column=2, value=f"=Daily_Headcount_Report!C{src_r}").font = NORMAL_FONT
    ws5.cell(row=r, column=3, value=f"=Daily_Headcount_Report!B{src_r}").font = NORMAL_FONT
    ws5.cell(row=r, column=4, value=f"=MAX(C{r}-B{r},0)").font = NORMAL_FONT
    ws5.cell(row=r, column=5, value=f"=ROUND(B{r}*$B$6*$B$5,0)").font = NORMAL_FONT
    ws5.cell(row=r, column=6, value=f"=D{r}+E{r}").font = NORMAL_FONT
    conv_ref = f"Recruitment_Funnel_Report!K{5+i}"
    ws5.cell(row=r, column=7, value=f"=IFERROR(F{r}/{conv_ref},F{r}/0.08)").font = NORMAL_FONT
    for col in range(1, 8):
        ws5.cell(row=r, column=col).border = BORDER

TOT_R = 12 + n_depts
ws5.cell(row=TOT_R, column=1, value="TOTAL").font = LABEL_FONT
for col_letter in ["B", "C", "D", "E", "F", "G"]:
    ws5.cell(row=TOT_R, column=ord(col_letter) - 64,
             value=f"=SUM({col_letter}12:{col_letter}{TOT_R-1})").font = LABEL_FONT

# fix B8 reference now that structure is final (average overall conversion % across depts)
ws5["B8"] = f"=AVERAGE(Recruitment_Funnel_Report!K5:K{4+n_depts})"

for j, w in enumerate([18, 18, 16, 20, 22, 18, 26], start=1):
    ws5.column_dimensions[get_column_letter(j)].width = w

wb.save(OUTFILE)
print(f"Saved workbook: {OUTFILE}")
